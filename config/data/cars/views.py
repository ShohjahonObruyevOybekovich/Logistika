# from rest_framework import generics
from decimal import Decimal

from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from openpyxl.styles import Alignment, Font
from openpyxl.workbook import Workbook
from rest_framework import generics, status
from rest_framework.exceptions import NotFound
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.generics import (
    ListAPIView,
    UpdateAPIView,
    DestroyAPIView,
    CreateAPIView, get_object_or_404, ListCreateAPIView, RetrieveUpdateDestroyAPIView, )
from rest_framework.generics import RetrieveAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .serializers import *
from ..finans.models import Logs
from ..flight.models import Flight


class CarCreateAPIView(CreateAPIView):
    queryset = Car.objects.all()
    serializer_class = CarCreateSerializer
    # permission_classes = (IsAuthenticated,)


class CarsListAPIView(ListAPIView):
    queryset = Car.objects.all().order_by("-created_at")
    serializer_class = CarListserializer
    # permission_classes = (IsAuthenticated,)

    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_fields = [
        "name",
        "number",
        "type_of_payment",
        "with_trailer",
        "fuel_type",
        "price_uzs",
        # "price_usd",
        "distance_travelled",
    ]
    ordering_fields = ["number"]
    search_fields = ["name"]


class CarsList_no_pg_APIView(ListAPIView):
    queryset = Car.objects.all().order_by("-created_at")
    serializer_class = CarListserializer
    # permission_classes = (IsAuthenticated,)

    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_fields = [
        "name",
        "number",
        "type_of_payment",
        "with_trailer",
        "fuel_type",
        "price_uzs",
        # "price_usd",
        "distance_travelled",
    ]
    ordering_fields = ["number"]
    search_fields = ["name"]

    def get_paginated_response(self, data):
        return Response(data)


class CarByIDAPIView(RetrieveAPIView):
    queryset = Car.objects.all().order_by("-created_at")
    serializer_class = CarListserializer
    permission_classes = (IsAuthenticated,)

    def get_object(self):
        try:
            obj = self.get_queryset().get(pk=self.kwargs.get("pk"))  # Use 'pk' here
            self.check_object_permissions(self.request, obj)
            return obj
        except Car.DoesNotExist:
            raise NotFound("Car not found.")


class DetailbyCarIDAPIView(ListAPIView):
    serializer_class = DetailCreateSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        car_id = self.kwargs.get("pk")
        queryset = Details.objects.filter(car_id=car_id).order_by("-created_at")

        if not queryset.exists():
            raise NotFound("No details found for the given car ID.")

        return queryset


class CarUpdateAPIView(UpdateAPIView):
    queryset = Car.objects.all()
    serializer_class = CarListserializer
    permission_classes = (IsAuthenticated,)


class ModelCarCreateAPIView(CreateAPIView):
    queryset = Model.objects.all()
    serializer_class = ModelSerializer
    permission_classes = (IsAuthenticated,)


class ModelCarListAPIView(ListAPIView):
    queryset = Model.objects.all().order_by("-created_at")
    serializer_class = ModelSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_fields = [
        'name'
    ]
    ordering_fields = ["name"]
    search_fields = ["name"]


class ModelCarList_no_pg_APIView(ListAPIView):
    queryset = Model.objects.all().order_by("-created_at")
    serializer_class = ModelSerializer
    permission_classes = (IsAuthenticated,)

    def get_paginated_response(self, data):
        return Response(data)


class ModelCarUpdateAPIView(UpdateAPIView):
    queryset = Model.objects.all()
    serializer_class = ModelSerializer
    permission_classes = (IsAuthenticated,)


class ModelCarDeleteAPIView(DestroyAPIView):
    queryset = Model.objects.all()
    serializer_class = ModelSerializer
    permission_classes = (IsAuthenticated,)


class DetailsCreateView(generics.ListCreateAPIView):
    queryset = Details.objects.all().order_by("-created_at")
    serializer_class = DetailCreateListSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        # The serializer will automatically call create method of ListSerializer
        serializer.save()


class DetailsView(ListAPIView):
    queryset = Details.objects.all().order_by("-created_at")
    serializer_class = DetailCreateSerializer
    # permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_fields = [
        "in_sklad",
        "id_detail",
        "name",
    ]
    ordering_fields = ["in_sklad"]
    search_fields = ["name"]


class BulkCreateUpdateAPIView(APIView):
    """
    Handle bulk creation and updating of `Details` objects.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        """
        Handle both creation and updating of Details objects.
        """
        data = request.data  # Expecting a list of objects
        if not isinstance(data, list) or not data:
            return Response(
                {"detail": "A non-empty list of objects is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        created_count = 0
        updated_count = 0
        errors = []

        for item in data:
            obj_id = item.get("id")  # Extract ID if present
            car_id = item.get("car")  # Extract car ID if present
            try:
                if obj_id:  # Update logic
                    detail = Details.objects.get(id=obj_id)
                    for field, value in item.items():
                        if field == "car" and car_id:
                            try:
                                car_instance = Car.objects.get(id=car_id)
                                setattr(detail, field, car_instance)
                            except Car.DoesNotExist:
                                errors.append({"id": obj_id, "detail": f"Car with id {car_id} not found"})
                                continue
                        elif field != "id":
                            setattr(detail, field, value)
                    detail.save()
                    updated_count += 1
                else:  # Create logic
                    if car_id:  # Resolve car foreign key for new object
                        try:
                            item["car"] = Car.objects.get(id=car_id)
                        except Car.DoesNotExist:
                            errors.append({"detail": f"Car with id {car_id} not found"})
                            continue
                    Details.objects.create(**item)
                    created_count += 1
            except Details.DoesNotExist:
                errors.append({"id": obj_id, "detail": "Object not found for update"})
            except Exception as e:
                errors.append({"id": obj_id, "detail": str(e)})

        # Prepare the response
        response_data = {
            "created_count": created_count,
            "updated_count": updated_count,
            "errors": errors,
        }

        return Response(response_data, status=status.HTTP_200_OK)


class BulkDeleteWithSellPriceAPIView(APIView):
    """
    Handle bulk deletion of `Details` objects with logging of a common sell price.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        """
        Delete multiple `Details` objects and log the sell price as income.
        """
        data = request.data  # Expecting a dict with 'ids' and 'sell_price'
        ids = data.get("id")  # List of IDs to delete

        sell_price = data.get("sell_price")
        sell_price_uzs = data.get("sell_price_uzs")
        sell_price_type = data.get("sell_price_type")



        # Validate input
        if not ids or not isinstance(ids, list):
            return Response(
                {"detail": "A non-empty list of IDs is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if sell_price and sell_price >= 0:
            # Log the income for the sell_price
            sell_price = Decimal(sell_price)
            Logs.objects.create(
                action="INCOME",
                amount_uzs=sell_price_uzs,
                amount=sell_price,
                amount_type=sell_price_type,
                kind="OTHER",
                comment=f"Детали проданы за {sell_price} {sell_price_type}.."
            )
        else:
            return Response(
                {"detail": "Sell price is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        deleted_count = 0
        errors = []

        for obj_id in ids:
            try:
                # Fetch the Details object by ID
                detail = Details.objects.get(id=obj_id)
                detail.delete()
                deleted_count += 1
            except Details.DoesNotExist:
                errors.append({"id": obj_id, "detail": "Details object not found."})
            except Exception as e:
                errors.append({"id": obj_id, "detail": str(e)})

        # Prepare the response
        response_data = {
            "deleted_count": deleted_count,
            "errors": errors,
        }

        return Response(response_data, status=status.HTTP_200_OK)


class DeleteCarAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        car_id = kwargs.get('uuid')  # Car UUID from URL
        sell_price = request.data.get('sell_price')

        if not sell_price :
            return Response(
                {"detail": "Sell price is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            sell_price_uzs = Decimal(sell_price)
        except Exception as e:
            return Response(
                {"detail": f"Invalid sell price: {e}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Get the car by its UUID
            car = Car.objects.get(id=car_id)
        except Car.DoesNotExist:
            return Response(
                {"detail": "Car not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"detail": f"Error fetching car: {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        try:
            # Log the sell price in the Logs model
            Logs.objects.create(
                action="INCOME",
                amount_uzs=sell_price_uzs,
                kind="OTHER",
                comment=f"Машина {car.name} - {car.number} продана за {sell_price_uzs} сум."
            )
        except Exception as e:
            return Response(
                {"detail": f"Error logging sell price: {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        try:
            # Delete the car
            car.delete()
        except Exception as e:
            return Response(
                {"detail": f"Error deleting car: {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response(
            {"detail": f"Car {car.name} successfully deleted and sell price logged."},
            status=status.HTTP_200_OK
        )


class DownloadCarInfoAPIView(APIView):
    # permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter("model", openapi.IN_QUERY, description="Filter by model ID", type=openapi.TYPE_STRING),
            openapi.Parameter("fuel_type", openapi.IN_QUERY, description="Filter by fuel type (DIESEL, GAS)",
                              type=openapi.TYPE_STRING),
        ],
        responses={200: "Excel file generated"}
    )
    def get(self, request):
        # Extract query parameters
        model_filter = request.GET.get("model")
        fuel_type_filter = request.GET.get("fuel_type")

        # Base queryset
        queryset = Car.objects.all()

        # Apply filters
        if model_filter:
            queryset = queryset.filter(model_id=model_filter)
        if fuel_type_filter:
            queryset = queryset.filter(fuel_type=fuel_type_filter)

        # Create an Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Car Info"

        # Define headers
        headers = [
            "Название", "Номер", "Модель", "Тип оплаты", "Срок лизинга",
            "С прицепом", "Тип топлива", "Цена (UZS)", "Пройденное расстояние",
            "Расстояние до замены масла", "Следующее расстояние до замены масла", "Номер прицепа", "Дата создания",
            "Дата обновления"
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.column_dimensions[cell.column_letter].width = 20

        # Write car data with fallback for missing fields
        for row_num, car in enumerate(queryset, start=2):
            sheet.cell(row=row_num, column=1).value = car.name or ""
            sheet.cell(row=row_num, column=2).value = car.number or ""
            sheet.cell(row=row_num, column=3).value = car.model.name if car.model else ""
            sheet.cell(row=row_num, column=4).value = car.type_of_payment or ""
            sheet.cell(row=row_num, column=5).value = car.leasing_period or ""
            sheet.cell(row=row_num, column=6).value = "Yes" if car.with_trailer else "No"
            sheet.cell(row=row_num, column=7).value = car.fuel_type or ""
            sheet.cell(row=row_num, column=8).value = car.price_uzs or ""
            sheet.cell(row=row_num, column=9).value = car.distance_travelled or ""
            sheet.cell(row=row_num, column=10).value = car.oil_recycle_distance or ""
            sheet.cell(row=row_num, column=11).value = car.next_oil_recycle_distance or ""
            sheet.cell(row=row_num, column=12).value = car.trailer_number or ""
            sheet.cell(row=row_num, column=13).value = car.created_at.strftime(
                '%Y-%m-%d %H:%M:%S') if car.created_at else ""
            sheet.cell(row=row_num, column=14).value = car.updated_at.strftime(
                '%Y-%m-%d %H:%M:%S') if car.updated_at else ""

        # Prepare the response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="Car_Info.xlsx"'

        # Save the workbook to the response
        workbook.save(response)
        return response


class FilteredCarDetailsExportToExcelView(ListAPIView):
    def get(self, request, *args, **kwargs):
        car_id = request.GET.get("car_id")  # Get the car_id from query parameters
        name_filter = request.GET.get("name")  # Get name filter from query parameters
        price_filter = request.GET.get("price_uzs")  # Get price filter from query parameters

        if not car_id:
            return HttpResponse("car_id is required", status=400)

        # Filter details by car_id
        details_queryset = Details.objects.filter(car_id=car_id)

        # Apply additional filters if they are provided
        if name_filter:
            details_queryset = details_queryset.filter(name__icontains=name_filter)

        if price_filter:
            try:
                price_filter = float(price_filter)  # Convert the price to a float for comparison
                details_queryset = details_queryset.filter(price_uzs=price_filter)
            except ValueError:
                return HttpResponse("Invalid price filter value.", status=400)

        if not details_queryset.exists():
            return HttpResponse("No details found for the specified car_id or filters.", status=404)

        # Create an Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Car Details"

        # Define headers
        headers = ["Название", "ИД", "Цена (UZS)", "В наличии", "Создано"]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.column_dimensions[cell.column_letter].width = 20

        # Write data rows
        for row_num, detail in enumerate(details_queryset, 2):
            sheet.cell(row=row_num, column=1).value = detail.name or ""
            sheet.cell(row=row_num, column=2).value = detail.id_detail or ""
            sheet.cell(row=row_num, column=3).value = detail.price_uzs or ""
            sheet.cell(row=row_num, column=4).value = "Да" if detail.in_sklad else "Нет"
            sheet.cell(row=row_num, column=5).value = detail.created_at.strftime(
                '%d-%m-%Y %H:%M') if detail.created_at else ""
            sheet.cell(row=row_num, column=6).value = detail.updated_at.strftime(
                '%d-%m-%Y %H:%M') if detail.updated_at else ""

        # Prepare the response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="Car_Details_{car_id}.xlsx"'
        workbook.save(response)
        return response


class CarDetailsExcelAPIView(APIView):
    """
    API View to generate an Excel file with car details and associated flights.
    """

    def get(self, request, pk, *args, **kwargs):
        # Get the car instance by pk or return 404 if not found
        car = get_object_or_404(Car, id=pk)

        # Fetch associated flights for the car
        flights_queryset = Flight.objects.filter(car=car)

        # Create an Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Car Details and Flights"

        # Define headers for car details
        car_headers = ["Название автомобиля", "Номер", "Модель", "Тип топлива", "Цена (UZS)", "Пройденное расстояние"]
        for col_num, header in enumerate(car_headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.column_dimensions[cell.column_letter].width = 20

        # Write car details
        car_data = [
            car.name,
            car.number,
            car.model.name if car.model else "",
            car.fuel_type,
            car.price_uzs or "",
            car.distance_travelled or 0
        ]
        for col_num, value in enumerate(car_data, 1):
            sheet.cell(row=2, column=col_num).value = value

        # Add headers for flights
        flights_start_row = 4
        flights_headers = ["Место отправления", "Место прибытия", "Дата отправления", "Дата прибытия", "Водитель", "Цена (UZS)", "Расстояние"]
        for col_num, header in enumerate(flights_headers, 1):
            cell = sheet.cell(row=flights_start_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.column_dimensions[cell.column_letter].width = 20

        # Write flights data
        for row_num, flight in enumerate(flights_queryset, flights_start_row + 1):
            flight_data = [
                "Туда" if flight.route == "GONE_TO" else "Туда и обратно" if flight.route == "BEEN_TO" else "",
                flight.region.name if flight.region else "",
                flight.departure_date.strftime('%d-%m-%Y') if flight.departure_date else "",
                flight.arrival_date.strftime('%d-%m-%Y') if flight.arrival_date else "",
                flight.driver.full_name if flight.driver else "",
                flight.price_uzs or "",
                flight.distance_travelled if hasattr(flight, 'distance_travelled') else ""
            ]
            for col_num, value in enumerate(flight_data, 1):
                sheet.cell(row=row_num, column=col_num).value = value

        # Prepare the HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="Car_Details_{car.number}.xlsx"'
        workbook.save(response)
        return response


class NotificationListApi(ListCreateAPIView):
    queryset = Notification.objects.all().order_by("-created_at")
    # permission_classes = [IsAuthenticated]
    serializer_class = Notificationserializer

    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_fields = [
        "is_read",
    ]
    ordering_fields = ["is_read"]
    search_fields = ["is_read"]

    def get_paginated_response(self, data):
        return Response(data)

class NotificationDetailsApi(RetrieveUpdateDestroyAPIView):
    queryset = Notification.objects.all().order_by("-created_at")
    permission_classes = [IsAuthenticated]
    serializer_class = Notificationserializer

class CarInfoApi(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        cars = Car.objects.all()
        total_price = cars.aggregate(total=Sum("price"))["total"] or 0
        total_price_uzs = cars.aggregate(total=Sum("price_uzs"))["total"] or 0

        return Response({
            "cars": cars.count(),
            "total_price_usd": total_price,
            "total_price_uzs": total_price_uzs
        })