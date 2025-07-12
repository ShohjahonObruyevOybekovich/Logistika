import openpyxl
from django.db.models import Sum, F, Case, When, FloatField, Q
from django.db.models.functions import TruncDay, TruncMonth
from django.http import HttpResponse
from django.utils.dateparse import parse_datetime
from django_filters.rest_framework import DjangoFilterBackend
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView, ListAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from data.finans.serializers import FinansListserializer, LogsFilter, FinansUserListserializer
from .models import Logs
from ..cars.models import Car
from ..flight.models import Flight
from ..gas.models import GasPurchase, GasSale, Gas_another_station
from ..salarka.models import Salarka, SalarkaAnotherStation


class Finans(ListCreateAPIView):
    queryset = Logs.objects.all().order_by("-created_at")
    serializer_class = FinansListserializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_fields = [
        "action",
        "amount_uzs",
        # "amount_usd",
        "car",
        "employee",
        "flight",
        "reason",
        "kind",
        "comment",
        "created_at",
    ]
    ordering_fields = ["action", "created_at"]
    search_fields = ["action", "reason", "employee", "flight", "created_at"]


class FinansList(ListAPIView):
    queryset = Logs.objects.all().order_by("-created_at")
    serializer_class = FinansUserListserializer
    # permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_class = LogsFilter
    ordering_fields = ["action", "created_at"]
    search_fields = ["action", "reason", "employee", "flight", "created_at"]

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Extract date filters from query parameters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Parse the dates if provided
        start_date = parse_datetime(start_date) if start_date else None
        end_date = parse_datetime(end_date) if end_date else None

        # Pass the filtered queryset and date range to the serializer context
        serializer = self.get_serializer(
            queryset,
            many=True,
            context={
                'filtered_queryset': queryset,
                'start_date': start_date,
                'end_date': end_date,
            }
        )

        # Paginate the results if necessary
        page = self.paginate_queryset(queryset)
        if page is not None:
            paginated_serializer = self.get_serializer(
                page,
                many=True,
                context={
                    'filtered_queryset': queryset,
                    'start_date': start_date,
                    'end_date': end_date,
                }
            )
            return self.get_paginated_response(paginated_serializer.data)

        return Response(serializer.data)


class FinansDetail(RetrieveUpdateDestroyAPIView):
    queryset = Logs.objects.all().order_by("-created_at")
    serializer_class = FinansListserializer
    permission_classes = [IsAuthenticated]



class FinansDriver(ListCreateAPIView):
    serializer_class = FinansListserializer
    permission_classes = [IsAuthenticated]  # Enforce authentication

    def get_queryset(self):
        driver_id = self.kwargs.get("pk")
        if driver_id:
            # Filter logs based on driver and specific kinds
            return Logs.objects.filter(
                employee__id=driver_id,
                kind__in=["PAY_SALARY", "FLIGHT_SALARY"]
            ).order_by("-created_at")

        return Logs.objects.none()


class FinansFlightExcel(APIView):
    def get(self, request, pk, *args, **kwargs):
        # Fetch the flight logs and related data
        flight_id = pk
        try:
            flight = Flight.objects.get(id=flight_id)
        except Flight.DoesNotExist:
            return HttpResponse("Flight not found", status=404)

        car = flight.car
        fuel_type = car.fuel_type.lower()  # Assuming fuel_type is either "salaka" or "gaz"

        # Fetch related logs
        logs = Logs.objects.filter(flight=flight, kind__in=["FLIGHT","FLIGHT_SALARY"]).order_by("-created_at")

        # Fetch purchases based on fuel type
        if fuel_type == "diesel":
            purchases = SalarkaAnotherStation.objects.filter(flight=flight).order_by("-created_at")
        elif fuel_type == "gas":
            purchases = list(
                GasSale.objects.filter(
                    car=car,
                    created_at__range=(flight.departure_date, flight.arrival_date)
                ).select_related("station")
            ) + list(
                Gas_another_station.objects.filter(
                    car=car,
                    created_at__range=(flight.departure_date, flight.arrival_date)
                )
            )
        else:
            return HttpResponse("Unsupported fuel type", status=400)

        # Create an Excel workbook
        wb = Workbook()
        ws_logs = wb.active
        ws_logs.title = "Flight Logs"

        # Calculate flight balance if flight is inactive
        flight_balance = None
        if flight.status == "INACTIVE":
            flight_balance = (
                flight.price_uzs
                - (
                    (flight.driver_expenses_uzs or 0)
                    + (flight.flight_expenses_uzs or 0)
                    + (flight.other_expenses_uzs or 0)
                )
            )

        # Define headers for logs
        headers_logs = [
            "Название",  # Name
            "Сумма (USD)",  # Amount (UZS)
            "Тип",  # Type
            "Комментарий",  # Comment
            "Начальная дата",  # Start Date
            "Конечная дата",  # End Date
            "Дата создания"  # Created At
        ]

        # Write headers for logs
        for col_num, header in enumerate(headers_logs, start=1):
            col_letter = get_column_letter(col_num)
            ws_logs[f"{col_letter}1"] = header

        # Write data rows for logs
        for row_num, log in enumerate(logs, start=2):
            ws_logs[f"A{row_num}"] = f"{flight.car.number} - {flight.region.name}"
            ws_logs[f"B{row_num}"] = log.amount_uzs
            ws_logs[f"C{row_num}"] = "Приход" if log.action == "INCOME" else "Расход"
            ws_logs[f"D{row_num}"] = f"{log.reason or ''} {log.comment or ''}"
            ws_logs[f"E{row_num}"] = flight.departure_date.strftime('%d-%m-%Y') if flight.departure_date else ""
            ws_logs[f"F{row_num}"] = flight.arrival_date.strftime('%d-%m-%Y') if flight.arrival_date else ""
            ws_logs[f"H{row_num}"] = log.created_at.strftime('%d-%m-%Y %H:%M')

        # Adjust column widths for logs
        for col_num, _ in enumerate(headers_logs, start=1):
            ws_logs.column_dimensions[get_column_letter(col_num)].width = 20

        # Add a sheet for purchases
        ws_purchases = wb.create_sheet(title="Purchases")
        headers_purchases = [
            "Название",  # Name
            "Количество",  # Quantity
            "Сумма (USD)",  # Amount (UZS)
            "Дата"  # Date
        ]

        # Write headers for purchases
        for col_num, header in enumerate(headers_purchases, start=1):
            col_letter = get_column_letter(col_num)
            ws_purchases[f"{col_letter}1"] = header

        # Write data rows for purchases
        for row_num, purchase in enumerate(purchases, start=2):
            if fuel_type == "salaka":
                ws_purchases[f"A{row_num}"] = f"{purchase.car.number} - {purchase.car.name}"
                ws_purchases[f"B{row_num}"] = purchase.volume
                ws_purchases[f"C{row_num}"] = f"{purchase.price} - {purchase.price_type}"
                ws_purchases[f"D{row_num}"] = purchase.created_at.strftime('%d-%m-%Y %H:%M')
            elif fuel_type == "gaz":
                name = purchase.station.name if hasattr(purchase, "station") else purchase.name
                ws_purchases[f"A{row_num}"] = name
                ws_purchases[f"B{row_num}"] = purchase.amount if hasattr(purchase, "amount") else purchase.purchased_volume
                ws_purchases[f"C{row_num}"] = f"{purchase.price} - {purchase.price_type}"
                ws_purchases[f"D{row_num}"] = purchase.created_at.strftime('%d-%m-%Y %H:%M')

        # Adjust column widths for purchases
        for col_num, _ in enumerate(headers_purchases, start=1):
            ws_purchases.column_dimensions[get_column_letter(col_num)].width = 20

        # Save the workbook to an in-memory response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="flight_logs_{flight_id}.xlsx"'
        wb.save(response)

        return response



class ExportLogsToExcelAPIView(APIView):
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter("action", openapi.IN_QUERY, description="Filter by action (INCOME, OUTCOME)",
                              type=openapi.TYPE_STRING),
            openapi.Parameter("amount_uzs", openapi.IN_QUERY, description="Filter by amount in UZS",
                              type=openapi.TYPE_NUMBER),
            openapi.Parameter("car", openapi.IN_QUERY, description="Filter by car ID", type=openapi.TYPE_STRING),
            openapi.Parameter("employee", openapi.IN_QUERY, description="Filter by employee ID",
                              type=openapi.TYPE_STRING),
            openapi.Parameter("flight", openapi.IN_QUERY, description="Filter by flight ID", type=openapi.TYPE_STRING),
            openapi.Parameter("reason", openapi.IN_QUERY, description="Filter by reason", type=openapi.TYPE_STRING),
            openapi.Parameter("kind", openapi.IN_QUERY, description="Filter by kind", type=openapi.TYPE_STRING),
            openapi.Parameter("comment", openapi.IN_QUERY, description="Filter by comment", type=openapi.TYPE_STRING),
            openapi.Parameter("start_date", openapi.IN_QUERY, description="Filter by start date (YYYY-MM-DD)",
                              type=openapi.TYPE_STRING),
            openapi.Parameter("end_date", openapi.IN_QUERY, description="Filter by end date (YYYY-MM-DD)",
                              type=openapi.TYPE_STRING),
        ],
        responses={200: "Excel file generated"}
    )
    def get(self, request):
        # Extract query parameters
        filters = {
            "action": request.GET.get("action"),
            "amount_uzs": request.GET.get("amount_uzs"),
            "car_id": request.GET.get("car"),
            "employee_id": request.GET.get("employee"),
            "flight_id": request.GET.get("flight"),
            "reason": request.GET.get("reason"),
            "kind": request.GET.get("kind"),
            "comment": request.GET.get("comment"),
        }
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        # Build the queryset with filters
        queryset = Logs.objects.all().order_by("-created_at")

        # Apply field-specific filters
        for field, value in filters.items():
            if value:
                filter_kwargs = {field: value}
                queryset = queryset.filter(**filter_kwargs)

        # Apply date range filter
        if start_date and end_date:
            queryset = queryset.filter(
                created_at__gte=start_date,
                created_at__lte=end_date
            )

        # Generate Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Logs"

        # Headers
        headers = ["Действие", "Сумма (USD)", "Машина", "Водитель", "Рейс", "Тип", "Причина", "Комментарий",
                   "Время создания"]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.column_dimensions[cell.column_letter].width = 20

        # Rows
        for row_num, log in enumerate(queryset, 2):
            sheet.cell(row=row_num, column=1).value = "Приход" if log.action == "INCOME" else "Расход"
            sheet.cell(row=row_num, column=2).value = log.amount_uzs
            sheet.cell(row=row_num, column=3).value = log.car.number if log.car else ""
            sheet.cell(row=row_num, column=4).value = log.employee.full_name if log.employee else ""
            sheet.cell(row=row_num, column=5).value = (
                f"{log.flight.car.number} - {log.flight.region.name}"
                if log.flight and log.flight.region and log.flight.car
                else ""
            )


            KIND_TO_RUSSIAN = {
                "OTHER": "Прочее",
                "FIX_CAR": "Ремонт автомобиля",
                "PAY_SALARY": "Зарплата",
                "SALARKA": "Солярка",
                "FLIGHT": "Рейс",
                "LEASING": "Лизинг",
                "BONUS": "Бонус",
                "BUY_CAR": "Купить машину."
            }

            # Populate the Excel cell with the Russian name
            sheet.cell(row=row_num, column=6).value = KIND_TO_RUSSIAN.get(log.kind, "")
            sheet.cell(row=row_num, column=7).value = log.reason if log.reason else ""
            sheet.cell(row=row_num, column=8).value = log.comment if log.comment else ""
            sheet.cell(row=row_num, column=9).value = log.created_at.strftime(
                '%d-%m-%Y   %H:%M') if log.created_at else ""

        # Response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="Logs_{filters.get("action", "All")}.xlsx"'
        workbook.save(response)
        return response


class FilteredIncomeOutcomeAPIView(APIView):
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter("year", openapi.IN_QUERY, description="Filter by year", type=openapi.TYPE_INTEGER),
            openapi.Parameter("month", openapi.IN_QUERY, description="Filter by month", type=openapi.TYPE_INTEGER),
            openapi.Parameter("day", openapi.IN_QUERY, description="Filter by day", type=openapi.TYPE_INTEGER),
            openapi.Parameter("start_date", openapi.IN_QUERY, description="Start date (YYYY-MM-DD)",
                              type=openapi.TYPE_STRING),
            openapi.Parameter("end_date", openapi.IN_QUERY, description="End date (YYYY-MM-DD)",
                              type=openapi.TYPE_STRING),
            openapi.Parameter("action", openapi.IN_QUERY, description="Filter by action (INCOME, OUTCOME)",
                              type=openapi.TYPE_STRING),
        ],
        responses={200: "Filtered income and outcome sums with chart data"}
    )
    def get(self, request):
        # Extract query parameters
        year = request.GET.get("year")
        month = request.GET.get("month")
        day = request.GET.get("day")
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        action = request.GET.get("action")  # Optional filter by action (INCOME, OUTCOME)

        # Base queryset
        queryset = Logs.objects.all().order_by("-created_at")

        # Apply year, month, and day filters
        if year:
            queryset = queryset.filter(created_at__year=year)
            if month:
                queryset = queryset.filter(created_at__month=month)
                if day:
                    queryset = queryset.filter(created_at__day=day)

        # Apply custom date range filter
        if start_date and end_date:
            queryset = queryset.filter(
                created_at__gte=start_date,
                created_at__lte=end_date
            )

        # Apply action filter (optional)
        if action in ["INCOME", "OUTCOME"]:
            queryset = queryset.filter(action=action)

        # Calculate income and outcome sums
        income_sum = queryset.filter(action="INCOME").aggregate(total_income=Sum("amount"))["total_income"] or 0
        outcome_sum = queryset.filter(action="OUTCOME").exclude(kind="BUY_CAR").aggregate(
            total_outcome=Sum("amount"))["total_outcome"] or 0

        # Determine grouping
        if start_date and end_date:
            group_by = TruncDay('created_at')
        elif year and month:
            group_by = TruncDay('created_at')
        elif year:
            group_by = TruncMonth('created_at')
        else:
            group_by = TruncMonth('created_at')

        # Generate chart data
        chart_data = (
            queryset
            .annotate(period=group_by)
            .values('period')
            .annotate(
                income=Sum(
                    Case(
                        When(action="INCOME", then=F('amount')),
                        default=0,
                        output_field=FloatField()
                    )
                ),
                outcome=Sum(
                    Case(
                        When(action="OUTCOME", then=F('amount')),
                        default=0,
                        output_field=FloatField()
                    )
                ),
            )
            .order_by('period')
        )

        # Prepare response data
        data = {
            "filters": {
                "year": year,
                "month": month,
                "day": day,
                "start_date": start_date,
                "end_date": end_date,
                "action": action,
            },
            "results": {
                "income_sum": income_sum,
                "outcome_sum": outcome_sum,
                "car_price": income_sum - outcome_sum,  # Net income
                "chart_data": list(chart_data)  # Grouped data for chart
            },
        }

        return Response(data)
